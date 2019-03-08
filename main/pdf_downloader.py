# ------------------------------------------------------------------
# FileName    : pdf_downloader.py
# Input       : List of urls's in main() routine
# Description : Crawl all links from given url and check
#               against regex to get relevant pdf links.
#               Downloaded pdf are store in pdf directory
# -------------------------------------------------------------------

import os
from urllib.request import urlopen, Request, urlparse
import re
import signal
import sys
import xlrd
from bs4 import BeautifulSoup
import urllib.parse
import logging
import ssl

# Logger configurations
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("Downloader")
workbook = None
excel_path = None
ssl_context = ssl._create_unverified_context()

headers = {
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 '
                  'Safari/537.11',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
    'Accept-Encoding': 'none',
    'Accept-Language': 'en-US,en;q=0.8',
    'Connection': 'keep-alive'}

regex = r"(investors|investorrelations|investor-relations)"
level = 0


def signal_handler(signal, frame):
    logger.warning('Shutting down PDF downloader engine....\n')
    if workbook:
        workbook.save(excel_path)
        workbook.close()
    sys.exit(0)


pdf_types = {'annual': 'AR', 'Corporate': 'AR', 'Transcripts': 'TR', 'Quarterly': 'QR',
             'Presentations': 'IP', 'Filings': 'FL'}

signal.signal(signal.SIGINT, signal_handler)


def get_pdf_type(text):
    result = ''
    for pdf_type in pdf_types.items():
        if re.search(pdf_type[0], text, re.IGNORECASE):
            result = pdf_type[1]
    if result:
        return result
    else:
        return 'OTH'


def get_file_size(num, suffix='B'):
    for unit in ['', 'Ki', 'Mi', 'Gi', 'Ti', 'Pi', 'Ei', 'Zi']:
        if abs(num) < 1024.0:
            return "%3.1f%s%s" % (num, unit, suffix)
        num /= 1024.0
    return "%.1f%s%s" % (num, 'Yi', suffix)


def get_pfd_urls_and_type(url, data, filtering=False):
    pdf_urls = []

    try:
        url_splits = url.split('/')
        pattern = url_splits[len(url_splits) - 1]
        match = re.search('.pdf', pattern, re.IGNORECASE)
        if match:
            logger.info("Fetching pdf url's for : {} || {}".format(str(data['CompanyId']).rstrip('0').rstrip('.'), url))
            logger.info("PDF download url's found : {}\n".format(1))
            doc_type = get_pdf_type(text=url)
            pdf_urls.append(url)
            return doc_type, pdf_urls

        url = convert_url(url)
        logger.info("\n")
        logger.info("Fetching pdf url's for : {} || {}".format(str(data['CompanyId']).rstrip('0').rstrip('.'), url))
        request = Request(url, headers=headers)
        response = urlopen(request, context=ssl_context)
        html_text = response.read()
        soup = BeautifulSoup(html_text, 'html.parser')
        title_text = soup.find('title').text
        doc_type = get_pdf_type(text=title_text)
        a_tags = soup.findAll('a', href=True)
        compile_regex = re.compile(regex, re.I)
        if a_tags:
            for a_tag in a_tags:
                if filtering:
                    match = compile_regex.search(a_tag['href'])
                    if a_tag['href'].endswith('.pdf') and 'javascript' not in a_tag['href'] and len(
                            a_tag['href']) > 20 and match is not None:
                        pdf_urls.append(get_domain(url) + '/' + a_tag['href'].replace(' ', '%20'))
                else:
                    if a_tag['href'].endswith('.pdf') and 'javascript' not in a_tag['href'] and len(
                            a_tag['href']) > 20:
                        pdf_urls.append(get_domain(url) + '/' + a_tag['href'].replace(' ', '%20'))
        pdf_urls = list(set(pdf_urls))
        logger.info("PDF download url's found : {}\n".format(len(pdf_urls)))
        return doc_type, pdf_urls
    except Exception as e:
        logger.exception(e)
        return 'Oth', []


def get_domain(url):
    parsed_url = urlparse(url)
    return parsed_url.scheme + '://' + parsed_url.netloc


# Rollback to previous version (without pdf type evaluation)
def get_pdf_meta(pdf_type, company_name, company_id, download_url):
    # pdf_title = str(company_id) + '_' + str(company_name).replace(' ', '_')
    pdf_title = str(company_id)
    for title in download_url.split('/'):
        if title.endswith('.pdf'):
            pdf_title = pdf_title + '_' + title.replace('%20', '_')
    # if pdf_type and pdf_type == 'OTH':
    #     doc_type = get_pdf_type(text=download_url)
    #     pdf_title = pdf_title + '_' + doc_type
    # for title in download_url.split('/'):
    #     if title.endswith('.pdf'):
    #         match = re.match(r'.*([1-3][0-9]{3})', title)
    #         if match is not None:
    #             year = match.group(1)
    #             pdf_title = pdf_title + '_' + year
    #         pdf_title = pdf_title + '.pdf'
    # pdf_folder = company_name.replace(' ', '_')
    pdf_folder = str(company_id) + '_' + company_name.replace(' ', '_')
    return pdf_title, pdf_folder


def download_pdf(download_url, pdf_title, pdf_folder):
    try:
        logger.info("PDF download url : {} ".format(download_url))
        logger.info("Downloading pdf ...")
        response = urlopen(download_url, context=ssl_context)
        if not os.path.exists('pdf'):
            os.makedirs('pdf')
        if not os.path.exists('pdf/{}'.format(pdf_folder)):
            os.makedirs('pdf/{}'.format(pdf_folder))
        if response.headers['content-length']:
            if int(response.headers['content-length']) > 0 and response.headers['content-type'].count(
                    'application/pdf'):
                file = open("pdf/{}/{}".format(pdf_folder, pdf_title), 'wb')
                file.write(response.read())
                file.close()
                logger.info(
                    "Downloaded PDF Name : {} || PDF-size : {} \n".format(pdf_title,
                                                                          get_file_size(int(
                                                                              response.headers['content-length']))))
        else:
            logger.error("{} url has zero content length".format(download_url))
    except Exception as e:
        logger.exception(e)


# read excel data and parse inputs for pdf extraction url's
def import_from_excel(excel_file, sheetname):
    try:
        rows = []
        header_list = []
        links_data = []
        wb = xlrd.open_workbook(excel_file)
        sheet = wb.sheet_by_index(0)
        logger.info("Reading file-name: {} || sheet-name: {}\n".format(excel_file, sheet.name))
        col = sheet.row(0)

        for header in col:
            header_list.append(header.value)
        for index in range(sheet.nrows):
            if not index == 0:
                rows.append(sheet.row_values(index))
        for row in rows:
            links_data.append(dict(zip(header_list, row)))
        wb.release_resources()
        del wb
        return links_data
    except Exception as e:
        logger.exception(e)


def convert_url(url):
    if url.startswith('http://'):
        return url
    if url.startswith('https://'):
        return url
    if url.startswith('www'):
        return 'http://' + url
    else:
        return 'http://' + url


def add_excel_headers(workbook, path, download_urls, data):
    sheet = workbook.create_sheet(title='pdf_urls')
    sheet.cell(1, 1).value = 'CompanyId'
    sheet.cell(1, 2).value = 'CompanyName'
    sheet.cell(1, 3).value = 'PDFUrl'

    for i in range(2, len(download_urls) + 2):
        sheet.cell(i, 1).value = data['CompanyId']
        sheet.cell(i, 2).value = data['CompanyName']
        sheet.cell(i, 3).value = download_urls[i - 2]
        workbook.save(path)
    # sheet.
    # sheet.cell(2, 1).value = data['CompanyId']
    # sheet.cell(2, 2).value = data['CompanyName']
    # sheet.cell(2, 3).value = download_urls[0]
    # sheet.cell(3, 1).value = data['CompanyId']
    # sheet.cell(3, 2).value = data['CompanyName']
    # sheet.cell(3, 3).value = download_urls[1]


# write pdf links to excel in another sheet
def export_to_excel(path, data):
    global workbook
    global excel_path
    try:
        excel_path = path
        workbook = load_workbook(excel_path)
        if 'pdf_urls' in workbook.sheetnames:
            workbook.remove_sheet(workbook.get_sheet_by_name('pdf_urls'))
            sheet = workbook.create_sheet(title='pdf_urls')
        else:
            sheet = workbook.create_sheet(title='pdf_urls')

        sheet.cell(1, 1).value = 'CompanyId'
        sheet.cell(1, 2).value = 'CompanyName'
        sheet.cell(1, 3).value = 'FilingsLink'
        sheet.cell(1, 4).value = 'PDFUrl'
        count = 2
        for i in data:
            pdf_type, pdf_urls = get_pfd_urls_and_type(url=i['FilingsLink'], data=i)

            for j in range(len(pdf_urls)):
                sheet.cell(j + count, 1).value = i['CompanyId']
                sheet.cell(j + count, 2).value = i['CompanyName']
                sheet.cell(j + count, 3).value = i['FilingsLink']
                sheet.cell(j + count, 4).value = pdf_urls[j]

                logger.info("Updating Row : {}...".format(sheet.max_row))
            count = sheet.max_row + 1
        workbook.save(path)
        workbook.close()
    except Exception as e:
        logger.exception(e)


def main(excel_file, level):
    try:
        if int(level) == 0:
            input_data = import_from_excel(excel_file=excel_file, sheetname='Export')  # list of excel records
            export_to_excel(path=excel_file, data=input_data)  # writing pdf url's to excel file

        if int(level) == 1:
            input_data = import_from_excel(excel_file=excel_file, sheetname='pdf_urls')

            for data in input_data:
                pdf_type, pdf_urls = get_pfd_urls_and_type(url=data['FilingsLink'], data=data, filtering=False)
                if len(pdf_urls):
                    for pdf_url in pdf_urls:
                        pdf_title, pdf_folder = get_pdf_meta(pdf_type=pdf_type, company_name=data['CompanyName'],
                                                             company_id=str(data['CompanyId']).rstrip('0').rstrip('.'),
                                                             download_url=pdf_url)
                        download_pdf(download_url=pdf_url, pdf_title=pdf_title, pdf_folder=pdf_folder)

    except Exception as e:
        logger.exception(e)

#
# def main(path):
#     try:
#         input_data = import_from_excel(path)  # list of excel records
#         export_to_excel(path=path, data=input_data)
#     except Exception as e:
#         logger.exception(e)
